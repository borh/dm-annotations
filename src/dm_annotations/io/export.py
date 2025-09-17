import csv
import re
import sqlite3
import subprocess
from itertools import chain
from pathlib import Path
from typing import Iterator, List, Union

import numpy as np
import orjson
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from dm_annotations.schemas import DmMatch


def export_dms(matches: Iterator[list[DmMatch]], file_name: str) -> None:
    """
    Serialize a sequence of DM‐match lists to CSV and Excel via Polars.

    :param matches: iterable of lists-of-dicts as from `extract_dm`
    :param file_name: target CSV path (will also write .xlsx)

    >>> from pathlib import Path
    >>> # one trivial match record
    >>> data = [[{"表現":"X","タイプ":"T","機能":"F","細分類":"S","position":0.1,"ジャンル":"G","title":"T1","sentence_id":0}]]
    >>> out = "test-dms.csv"
    >>> export_dms(data, out)
    >>> Path(out).exists()
    True
    >>> Path(out).unlink(); Path("test-dms.xlsx").unlink()
    """
    # Flatten and include the actual matched text instead of dropping the span entirely
    raw = list(chain.from_iterable(matches))
    # Ensure known DmMatch keys are always present, even if raw is empty or missing fields
    base_fields = [
        "表現",
        "タイプ",
        "機能",
        "細分類",
        "position",
        "ジャンル",
        "title",
        "sentence_id",
        "section",
    ]
    all_keys = sorted(
        set(base_fields) | {k for rec in raw for k in rec.keys() if k != "span"}
    )
    if "span_text" not in all_keys:
        all_keys.append("span_text")
    # Build uniform records with all keys present
    records = []
    for rec in raw:
        row = {}
        for k in all_keys:
            if k == "span_text":
                row[k] = rec["span"].text if "span" in rec else None
            else:
                row[k] = rec.get(k, None)
        records.append(row)
    # Always pass the schema (column names) to polars so it uses them consistently
    df = pl.DataFrame(records, schema=all_keys)
    df.write_csv(file_name)
    try:
        df.write_excel(
            f"{Path(file_name).stem}.xlsx",
            header_format={"bold": True},
            autofit=True,
            freeze_panes=(1, 0),
        )
    except Exception as e:
        print(f"Skipping Excel export: {e}")


def export_count(file_name: str) -> None:
    """
    Count DM‐matches by type/genre/function/subclass/expression,
    writing `<stem>-counts.csv` and `-counts.xlsx`.

    :param file_name: CSV path previously written by `export_dms`

    >>> from pathlib import Path; import os
    >>> out = "dummy.csv"
    >>> Path(out).write_text("タイプ,ジャンル,機能,細分類,表現\\n")  # no data
    >>> export_count(out)
    >>> Path("dummy-counts.csv").exists()
    True
    >>> os.remove(out); os.remove("dummy-counts.csv"); os.remove("dummy-counts.xlsx")
    """
    df = pl.read_csv(file_name)
    result_df = (
        df.group_by(["タイプ", "ジャンル", "機能", "細分類", "表現"])
        .count()
        .rename({"count": "頻度"})
        .sort(["タイプ", "ジャンル", "頻度"], descending=[False, False, True])
    )
    file_path = Path(file_name)
    stem = file_path.stem
    # write counts files beside the input CSV
    csv_out = file_path.parent / f"{stem}-counts.csv"
    xlsx_out = file_path.parent / f"{stem}-counts.xlsx"
    result_df.write_csv(csv_out)
    result_df.write_excel(
        xlsx_out,
        header_format={"bold": True},
        autofit=True,
        freeze_panes=(1, 0),
    )


def export_surface_forms(
    dm_csv: Union[str, Path],
    out_dir: Union[str, Path],
    corpus_jsonl: Union[Path, None] = None,
    include_samples: bool = False,
) -> None:
    """
    Optimized version using Polars' lazy API and streaming for memory efficiency.
    """
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    # make subfolders for sentence‐final and connectives outputs
    sf_dir = out_path / "sf"
    conn_dir = out_path / "conn"
    sf_dir.mkdir(parents=True, exist_ok=True)
    conn_dir.mkdir(parents=True, exist_ok=True)

    sid_to_text = {}
    if include_samples and corpus_jsonl:
        sid = 0
        with open(corpus_jsonl, "r", encoding="utf-8") as f:
            for line in f:
                record = orjson.loads(line)
                for sent in record["sentences"]:
                    sid_to_text[sid] = sent
                    sid += 1
    print(f"Read in {len(sid_to_text)} sentences from corpus.")

    def safe(expr: str) -> str:
        return re.sub(r"[^\w\-]", "_", expr)

    # Optimized surface formatter using native operations
    def format_surfaces_fast(texts: List[str], n: int = 10) -> str:
        if not texts:
            return ""

        # Limit processing to a sample for performance within map_elements
        # This sampling is applied per group after initial aggregations.
        sample_texts = texts[:1000]

        unique, counts = np.unique(sample_texts, return_counts=True)
        sorted_idx = np.argsort(-counts)

        parts = []
        total_shown_in_sample = 0
        for i in range(min(n, len(unique))):
            idx = sorted_idx[i]
            parts.append(f"{unique[idx]}: {counts[idx]}")
            total_shown_in_sample += counts[idx]

        num_in_sample_not_top_n = len(sample_texts) - total_shown_in_sample
        num_outside_sample = len(texts) - len(sample_texts)
        rest_count = num_in_sample_not_top_n + num_outside_sample

        if rest_count > 0:
            parts.append(f"他({rest_count}): {rest_count}")

        return "; ".join(parts)

    # ---------------------------------------------
    # Disk-backed streaming via SQLite to bound memory
    # ---------------------------------------------
    db_path = out_path / "dm_cache.db"
    conn = sqlite3.connect(db_path.as_posix())
    cur = conn.cursor()
    # speed up bulk‐load from Python metadata
    cur.executescript("""
        PRAGMA journal_mode = OFF;
        PRAGMA synchronous  = OFF;
        PRAGMA temp_store   = MEMORY;
    """)
    # create an empty table; we will bulk‐import via sqlite3 CLI
    cur.execute(
        "CREATE TABLE dms(sentence_id INTEGER, type TEXT, expr TEXT, text TEXT)"
    )
    # commit the empty table & close Python connection so sqlite3 CLI can import
    conn.commit()
    conn.close()

    # 1) dump a tiny CSV with exactly four columns in the right order
    tmp = out_path / "dms_tmp.csv"
    with (
        open(dm_csv, newline="", encoding="utf-8") as src,
        open(tmp, "w", newline="", encoding="utf-8") as dst,
    ):
        rdr = csv.DictReader(src)
        wtr = csv.writer(dst)
        for rec in tqdm(rdr):
            wtr.writerow(
                (
                    rec["sentence_id"],
                    rec["タイプ"],
                    rec["表現"],
                    rec["span_text"],
                )
            )

    # 2) call sqlite3 cli to import in one shot
    print(f"Bulk loading with sqlite from {tmp}...")
    subprocess.run(
        ["sqlite3", db_path.as_posix()],
        input=f".mode csv\n.import {tmp.as_posix()} dms\n",
        text=True,
        check=True,
    )
    # comment out the unlink if you want to keep the temp file around:
    # tmp.unlink()
    print(f"[debug] temporary CSV retained at {tmp}")

    # 3) reopen Python connection and build the index
    conn = sqlite3.connect(db_path.as_posix())
    cur = conn.cursor()
    print("Creating index...")
    cur.execute("CREATE INDEX idx_sentence ON dms(sentence_id)")
    # add partial indexes on expr for each type and collect stats
    cur.executescript("""
      CREATE INDEX IF NOT EXISTS idx_sf_expr 
        ON dms(expr) WHERE type='文末表現';
      CREATE INDEX IF NOT EXISTS idx_conn_expr 
        ON dms(expr) WHERE type='接続表現';
      ANALYZE;
    """)
    print("Index created.")

    # 1) SF → Conn via one GROUP BY query, streaming per‐pattern CSV
    sf_sql = """
WITH
  sf AS (
    SELECT sentence_id, expr AS sf_expr
      FROM dms
     WHERE type = '文末表現'
  ),
  conn AS (
    SELECT sentence_id, expr AS conn_expr, text AS conn_text
      FROM dms
     WHERE type = '接続表現'
  ),
  matched AS (
    SELECT sf.sf_expr, conn.conn_expr, conn.conn_text, sf.sentence_id AS sid
      FROM sf
      JOIN conn USING(sentence_id)
  ),
  surface_counts AS (
    SELECT
      sf_expr, conn_expr, conn_text,
      COUNT(*) AS surf_freq
    FROM matched
    GROUP BY sf_expr, conn_expr, conn_text
  ),
  sample_sids AS (
    SELECT sf_expr, conn_expr,
           GROUP_CONCAT(sid, '||') AS sids
      FROM (
        SELECT sf_expr, conn_expr, sid,
               ROW_NUMBER() OVER (
                 PARTITION BY sf_expr, conn_expr
                 ORDER BY RANDOM()
               ) AS rn
          FROM matched
      )
     WHERE rn <= 5
     GROUP BY sf_expr, conn_expr
  )
SELECT
  sc.sf_expr,
  sc.conn_expr,
  SUM(sc.surf_freq)                                AS freq,
  GROUP_CONCAT(sc.conn_text || ': ' || sc.surf_freq, '; ') AS surface_forms,
  sample_sids.sids                                 AS sids
FROM surface_counts AS sc
LEFT JOIN sample_sids USING(sf_expr, conn_expr)
GROUP BY sc.sf_expr, sc.conn_expr
ORDER BY sc.sf_expr, freq DESC
"""
    cur.execute(sf_sql)
    current_sf = None
    writer = None
    csv_file = None
    print("Exporting sfs")
    for sf_expr, conn_expr, freq, surface_forms, sids in tqdm(cur):
        if sf_expr != current_sf:
            if writer and csv_file:
                csv_file.close()
            current_sf = sf_expr
            csv_file = open(
                sf_dir / f"sf_{safe(str(sf_expr))}.csv",
                "w",
                newline="",
                encoding="utf-8",
            )
            writer = csv.writer(csv_file)
            header = ["connective", "freq", "surface_forms"]
            if include_samples:
                header += [f"sample_{i}" for i in range(1, 6)]
            if writer:
                writer.writerow(header)

        # parse & sort ALL forms by descending freq, then take top10 + “他” summary
        raw_parts = (surface_forms or "").split("; ")
        items: list[tuple[str, int]] = []
        for part in raw_parts:
            if not part:
                continue
            name, cnt_str = part.rsplit(": ", 1)
            items.append((name, int(cnt_str)))
        items.sort(key=lambda x: -x[1])
        if len(items) > 10:
            top_items = items[:10]
            rest = items[10:]
            rest_types = len(rest)
            rest_freq = sum(cnt for _, cnt in rest)
            display = [f"{n}: {c}" for n, c in top_items]
            display.append(f"他({rest_types}): {rest_freq}")
        else:
            display = [f"{n}: {c}" for n, c in items]
        surface_display = "; ".join(display)
        row = [conn_expr, freq, surface_display]
        if include_samples:
            sample_ids = [int(x) for x in sids.split("||")[:5]] if sids else []
            row += [sid_to_text.get(sid, "") for sid in sample_ids]
        if writer:
            writer.writerow(row)
    if writer and csv_file:
        csv_file.close()

    # 2) Conn → SF via one GROUP BY query
    conn_sql = """
WITH
  conn AS (
    SELECT sentence_id, expr AS conn_expr
      FROM dms
     WHERE type = '接続表現'
  ),
  sf AS (
    SELECT sentence_id, expr AS sf_expr, text AS sf_text
      FROM dms
     WHERE type = '文末表現'
  ),
  matched AS (
    SELECT conn.conn_expr, sf.sf_expr, sf.sf_text AS conn_text, conn.sentence_id AS sid
      FROM conn
      JOIN sf USING(sentence_id)
  ),
  surface_counts AS (
    SELECT
      conn_expr, sf_expr, conn_text,
      COUNT(*) AS surf_freq
    FROM matched
    GROUP BY conn_expr, sf_expr, conn_text
  ),
  sample_sids AS (
    SELECT conn_expr AS sf_expr, sf_expr AS conn_expr,
           GROUP_CONCAT(sid, '||') AS sids
      FROM (
        SELECT conn_expr AS sf_expr, sf_expr AS conn_expr, sid,
               ROW_NUMBER() OVER (
                 PARTITION BY conn_expr, sf_expr
                 ORDER BY RANDOM()
               ) AS rn
          FROM matched
      )
     WHERE rn <= 5
     GROUP BY sf_expr, conn_expr
  )
SELECT
  sc.conn_expr      AS conn_expr,
  sc.sf_expr        AS sf_expr,
  SUM(sc.surf_freq) AS freq,
  GROUP_CONCAT(sc.conn_text || ': ' || sc.surf_freq, '; ') AS surface_forms,
  sample_sids.sids  AS sids
FROM surface_counts AS sc
LEFT JOIN sample_sids ON sc.conn_expr = sample_sids.conn_expr
                      AND sc.sf_expr   = sample_sids.sf_expr
GROUP BY sc.conn_expr, sc.sf_expr
ORDER BY sc.conn_expr, freq DESC
"""
    cur.execute(conn_sql)
    current_conn = None
    writer = None
    csv_file = None
    print("Exporting conns")
    for conn_expr, sf_expr, freq, surface_forms, sids in tqdm(cur):
        if conn_expr != current_conn:
            if writer and csv_file:
                csv_file.close()
            current_conn = conn_expr
            csv_file = open(
                conn_dir / f"conn_{safe(str(conn_expr))}.csv",
                "w",
                newline="",
                encoding="utf-8",
            )
            writer = csv.writer(csv_file)
            header = ["sf_pattern", "freq", "surface_forms"]
            if include_samples:
                header += [f"sample_{i}" for i in range(1, 6)]
            if writer:
                writer.writerow(header)

        # parse & sort ALL forms by descending freq, then take top10 + “他” summary
        raw_parts = (surface_forms or "").split("; ")
        items: list[tuple[str, int]] = []
        for part in raw_parts:
            if not part:
                continue
            name, cnt_str = part.rsplit(": ", 1)
            items.append((name, int(cnt_str)))
        items.sort(key=lambda x: -x[1])
        if len(items) > 10:
            top_items = items[:10]
            rest = items[10:]
            rest_types = len(rest)
            rest_freq = sum(cnt for _, cnt in rest)
            display = [f"{n}: {c}" for n, c in top_items]
            display.append(f"他({rest_types}): {rest_freq}")
        else:
            display = [f"{n}: {c}" for n, c in items]
        surface_display = "; ".join(display)
        row = [sf_expr, freq, surface_display]
        if include_samples:
            sample_ids = [int(x) for x in sids.split("||")[:5]] if sids else []
            row += [sid_to_text.get(sid, "") for sid in sample_ids]
        if writer:
            writer.writerow(row)
    if writer and csv_file:
        csv_file.close()
    conn.close()

    # --- write a one‐sheet Excel next to each group‐CSV ---
    import csv as _csv

    from openpyxl import Workbook as _WB

    def _write_sheet_from_csv(csv_path: Path) -> None:
        """Read csv_path via Python csv and write a one‐sheet .xlsx alongside it."""
        wb1 = _WB()
        ws1 = wb1.active
        ws1.title = csv_path.stem[:31]
        with open(csv_path, newline="", encoding="utf-8") as f_in:
            reader = _csv.reader(f_in)
            for i, row in enumerate(reader):
                if i == 0:
                    # header row
                    ws1.append(row)
                else:
                    # coerce freq col (index 1) to int if possible
                    try:
                        row[1] = int(row[1])
                    except Exception:
                        try:
                            row[1] = float(row[1])
                        except Exception:
                            pass
                    ws1.append(row)

        # custom formatting: bold header, freeze it, autofit columns
        for cell in ws1[1]:
            cell.font = Font(bold=True)
        ws1.freeze_panes = "A2"
        for col in ws1.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max_length + 2

        wb1.save(csv_path.with_suffix(".xlsx"))

    for csv_path in sorted(sf_dir.glob("*.csv")):
        _write_sheet_from_csv(csv_path)
    for csv_path in sorted(conn_dir.glob("*.csv")):
        _write_sheet_from_csv(csv_path)

    # --- now create the combined Excel (all_results.xlsx) as before ---
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # 0) summary sheet: include the two surface‐forms columns
    ws_all = wb.create_sheet(title="collocations", index=0)
    ws_all.append(
        ["sf_expr", "sf_surface_forms", "conn_expr", "conn_surface_forms", "freq"]
    )
    for sf_csv in sorted(sf_dir.glob("*.csv")):
        sf_expr = sf_csv.stem.removeprefix("sf_")
        with open(sf_csv, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)  # skip header
            for row in reader:
                # row[2] is actually the CONN surface_forms in sf_*.csv
                conn_expr, freq, conn_forms, *rest = row
                try:
                    freq_val = int(freq)
                except ValueError:
                    freq_val = freq
                # we’ll pull the SF surface_forms from conn_*.csv
                conn_csv = conn_dir / f"conn_{safe(str(conn_expr))}.csv"
                sf_forms = ""
                with open(conn_csv, newline="", encoding="utf-8") as cf2:
                    cr = csv.reader(cf2)
                    next(cr, None)
                    for r2 in cr:
                        if r2[0] == sf_expr:
                            _, _, sf_forms, *_ = r2
                            break
                ws_all.append([sf_expr, sf_forms, conn_expr, conn_forms, freq_val])

    # also emit a standalone collocations.csv
    coll_path = out_path / "collocations.csv"
    with open(coll_path, "w", newline="", encoding="utf-8") as cf:
        cw = csv.writer(cf)
        # new header: include SF‐ and Conn‐surface‐forms
        cw.writerow(
            ["sf_expr", "sf_surface_forms", "conn_expr", "conn_surface_forms", "freq"]
        )
        for sf_csv in sorted(sf_dir.glob("*.csv")):
            sf_expr = sf_csv.stem.removeprefix("sf_")
            with open(sf_csv, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    # row[2] is actually the CONN surface_forms in sf_*.csv
                    conn_expr, freq, conn_forms, *rest = row
                    # cast freq
                    try:
                        freq_val = int(freq)
                    except ValueError:
                        freq_val = freq
                    # we’ll pull the SF surface_forms from conn_*.csv
                    conn_csv = conn_dir / f"conn_{safe(str(conn_expr))}.csv"
                    sf_forms = ""
                    with open(conn_csv, newline="", encoding="utf-8") as cf2:
                        cr = csv.reader(cf2)
                        next(cr, None)
                        for r2 in cr:
                            if r2[0] == sf_expr:
                                _, _, sf_forms, *_ = r2
                                break
                    cw.writerow([sf_expr, sf_forms, conn_expr, conn_forms, freq_val])

    # include all group‐CSV sheets from sf/ and conn/
    for csv_path in sorted(sf_dir.glob("*.csv")) + sorted(conn_dir.glob("*.csv")):
        ws = wb.create_sheet(title=csv_path.stem[:31])
        with open(csv_path, newline="", encoding="utf-8") as _f:
            reader = csv.reader(_f)
            for i, row in enumerate(reader):
                if i == 0:
                    ws.append(row)
                else:
                    # coerce freq col (index 1) to int if possible
                    try:
                        row[1] = int(row[1])
                    except Exception:
                        try:
                            row[1] = float(row[1])
                        except Exception:
                            pass
                    ws.append(row)

    # custom formatting for every sheet in collocations.xlsx
    for ws in wb.worksheets:
        # bold the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # freeze the top row
        ws.freeze_panes = "A2"
        # autofit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

    # save the full multi‐sheet workbook under the new name
    wb.save(out_path / "all_results.xlsx")

    # now that collocations.csv exists, create its single‐sheet .xlsx
    _write_sheet_from_csv(coll_path)

    print("Processing complete. Output files are in:", out_path)
